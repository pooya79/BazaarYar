from __future__ import annotations

import base64
import csv
import io
import json
import mimetypes
import re
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from uuid import UUID, uuid4
from xml.etree import ElementTree as ET

from fastapi import HTTPException, UploadFile
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from server.core.config import get_settings
from server.db.models import Attachment

from .schemas import AttachmentMediaType, StoredAttachment

_IMAGE_MIME_PREFIX = "image/"
_PDF_MIME = "application/pdf"
_TEXT_MIME_TYPES = {
    "text/plain",
    "text/csv",
    "application/csv",
    "text/tab-separated-values",
}
_SPREADSHEET_MIME_TYPES = {
    "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
_ALLOWED_MIME_TYPES = {
    _PDF_MIME,
    *_TEXT_MIME_TYPES,
    *_SPREADSHEET_MIME_TYPES,
}
_ALLOWED_EXTENSIONS = {
    ".txt",
    ".csv",
    ".tsv",
    ".pdf",
    ".xlsx",
    ".xls",
    ".png",
    ".jpg",
    ".jpeg",
    ".webp",
    ".gif",
}
_TEXT_PREVIEW_LIMIT = 2_000
_TABLE_PREVIEW_ROW_LIMIT = 5
_TABLE_PREVIEW_COL_LIMIT = 20


def _storage_root() -> Path:
    settings = get_settings()
    root = Path(settings.upload_storage_dir)
    if not root.is_absolute():
        project_root = Path(__file__).resolve().parents[3]
        root = project_root / root
    root.mkdir(parents=True, exist_ok=True)
    (root / "files").mkdir(parents=True, exist_ok=True)
    return root


def _metadata_path(file_id: str) -> Path:
    return _storage_root() / "metadata" / f"{file_id}.json"


def list_legacy_metadata_attachments() -> list[StoredAttachment]:
    metadata_dir = _storage_root() / "metadata"
    if not metadata_dir.exists():
        return []

    items: list[StoredAttachment] = []
    for path in metadata_dir.glob("*.json"):
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
            items.append(StoredAttachment.model_validate(payload))
        except Exception:
            continue
    return items


def _normalize_filename(filename: str) -> str:
    cleaned = re.sub(r"[^\w.\- ]+", "_", filename).strip()
    return cleaned or "upload"


def _content_type_for(upload: UploadFile, filename: str) -> str:
    from_upload = (upload.content_type or "").strip().lower()
    if from_upload:
        return from_upload
    guessed, _ = mimetypes.guess_type(filename)
    return (guessed or "application/octet-stream").lower()


def _file_extension(filename: str) -> str:
    return Path(filename).suffix.lower()


def _is_allowed(content_type: str, extension: str) -> bool:
    if content_type.startswith(_IMAGE_MIME_PREFIX):
        return True
    if content_type in _ALLOWED_MIME_TYPES:
        return True
    return extension in _ALLOWED_EXTENSIONS


def _media_type(content_type: str, extension: str) -> AttachmentMediaType:
    if content_type.startswith(_IMAGE_MIME_PREFIX) or extension in {
        ".png",
        ".jpg",
        ".jpeg",
        ".webp",
        ".gif",
    }:
        return "image"
    if content_type == _PDF_MIME or extension == ".pdf":
        return "pdf"
    if content_type in _TEXT_MIME_TYPES or extension in {".txt", ".csv", ".tsv"}:
        return "text"
    if content_type in _SPREADSHEET_MIME_TYPES or extension in {".xlsx", ".xls"}:
        return "spreadsheet"
    return "binary"


def infer_attachment_media_type(content_type: str, filename: str) -> AttachmentMediaType:
    return _media_type((content_type or "").lower(), _file_extension(filename))


def _clip_text(value: str, max_chars: int = _TEXT_PREVIEW_LIMIT) -> str:
    if len(value) <= max_chars:
        return value
    return f"{value[:max_chars].rstrip()}\n\n[truncated]"


def _format_table_preview(rows: list[list[str]]) -> str:
    cleaned_rows: list[list[str]] = []
    for row in rows[:_TABLE_PREVIEW_ROW_LIMIT]:
        cleaned = [cell.strip() for cell in row[:_TABLE_PREVIEW_COL_LIMIT]]
        if any(cleaned):
            cleaned_rows.append(cleaned)

    if not cleaned_rows:
        return ""

    rendered = ["\t".join(row) for row in cleaned_rows]
    return _clip_text("\n".join(rendered))


def _extract_text_preview(data: bytes, content_type: str, extension: str) -> tuple[str | None, str | None]:
    if content_type.startswith(_IMAGE_MIME_PREFIX) or extension in {
        ".png",
        ".jpg",
        ".jpeg",
        ".webp",
        ".gif",
    }:
        return None, "Image preview is available in UI; OCR is not enabled yet."

    if content_type == _PDF_MIME or extension == ".pdf":
        return _extract_pdf_preview(data)

    if extension == ".xlsx" or content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return _extract_xlsx_preview(data)

    if extension == ".xls" or content_type == "application/vnd.ms-excel":
        return (
            None,
            "Binary .xls parsing is not enabled yet. Use .xlsx for text extraction.",
        )

    if extension in {".csv", ".tsv"} or content_type in {
        "text/csv",
        "application/csv",
        "text/tab-separated-values",
    }:
        delimiter = "\t" if extension == ".tsv" or content_type == "text/tab-separated-values" else ","
        return _extract_csv_like_preview(data, delimiter=delimiter)

    if extension == ".txt" or content_type == "text/plain":
        try:
            return _clip_text(data.decode("utf-8", errors="replace")), None
        except Exception as exc:  # pragma: no cover - defensive
            return None, f"Could not parse text file: {exc}"

    return None, "No text extraction configured for this file type."


def _extract_csv_like_preview(data: bytes, *, delimiter: str) -> tuple[str | None, str | None]:
    try:
        decoded = data.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(decoded), delimiter=delimiter)
        rows: list[list[str]] = []
        for row in reader:
            rows.append([str(cell) for cell in row])
            if len(rows) >= _TABLE_PREVIEW_ROW_LIMIT:
                break
        preview = _format_table_preview(rows)
        if not preview:
            return None, "File had no non-empty rows."
        return preview, None
    except Exception as exc:
        return None, f"Could not parse table file: {exc}"


def _extract_pdf_preview(data: bytes) -> tuple[str | None, str | None]:
    try:
        from pypdf import PdfReader
    except Exception:
        return None, "PDF text extraction dependency is unavailable (install pypdf)."

    try:
        reader = PdfReader(io.BytesIO(data))
        chunks: list[str] = []
        for page in reader.pages[:5]:
            text = (page.extract_text() or "").strip()
            if text:
                chunks.append(text)
            if len("\n".join(chunks)) >= _TEXT_PREVIEW_LIMIT:
                break
        if not chunks:
            return None, "No extractable text found in PDF."
        return _clip_text("\n\n".join(chunks)), None
    except Exception as exc:
        return None, f"Could not parse PDF: {exc}"


def _extract_xlsx_preview(data: bytes) -> tuple[str | None, str | None]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return _extract_xlsx_preview_stdlib(data)

    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.worksheets[0] if workbook.worksheets else None
        if sheet is None:
            return None, "Workbook has no worksheets."
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            rows.append([("" if cell is None else str(cell)) for cell in row])
            if len(rows) >= _TABLE_PREVIEW_ROW_LIMIT:
                break
        preview = _format_table_preview(rows)
        if not preview:
            return None, "Worksheet had no visible values."
        return preview, None
    except Exception as exc:
        return None, f"Could not parse xlsx workbook: {exc}"


def _extract_xlsx_preview_stdlib(data: bytes) -> tuple[str | None, str | None]:
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            shared_strings: list[str] = []
            if "xl/sharedStrings.xml" in zf.namelist():
                shared_root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
                for item in shared_root.findall(".//x:si", ns):
                    text_parts = [node.text or "" for node in item.findall(".//x:t", ns)]
                    shared_strings.append("".join(text_parts))

            sheet_names = sorted(
                name
                for name in zf.namelist()
                if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
            )
            if not sheet_names:
                return None, "Workbook has no worksheets."

            sheet_root = ET.fromstring(zf.read(sheet_names[0]))
            rows: list[list[str]] = []
            for row in sheet_root.findall(".//x:sheetData/x:row", ns):
                out_row: list[str] = []
                for cell in row.findall("x:c", ns):
                    cell_type = cell.attrib.get("t")
                    inline_node = cell.find("x:is/x:t", ns)
                    value_node = cell.find("x:v", ns)
                    if inline_node is not None and inline_node.text:
                        out_row.append(inline_node.text)
                        continue
                    if value_node is None or value_node.text is None:
                        out_row.append("")
                        continue
                    raw_value = value_node.text
                    if cell_type == "s":
                        try:
                            out_row.append(shared_strings[int(raw_value)])
                        except Exception:
                            out_row.append(raw_value)
                    else:
                        out_row.append(raw_value)
                rows.append(out_row)
                if len(rows) >= _TABLE_PREVIEW_ROW_LIMIT:
                    break

            preview = _format_table_preview(rows)
            if not preview:
                return None, "Worksheet had no visible values."
            return preview, None
    except Exception as exc:
        return None, f"Could not parse xlsx workbook: {exc}"


async def _read_file_limited(upload: UploadFile, *, max_size: int) -> bytes:
    total = 0
    chunks: list[bytes] = []
    while True:
        chunk = await upload.read(1024 * 1024)
        if not chunk:
            break
        total += len(chunk)
        if total > max_size:
            raise HTTPException(
                status_code=413,
                detail=f"File '{upload.filename or 'upload'}' exceeds max size of {max_size} bytes.",
            )
        chunks.append(chunk)
    return b"".join(chunks)


async def store_uploaded_file(upload: UploadFile) -> StoredAttachment:
    settings = get_settings()

    if not upload.filename:
        raise HTTPException(status_code=400, detail="Uploaded file is missing a filename.")

    normalized_name = _normalize_filename(upload.filename)
    extension = _file_extension(normalized_name)
    content_type = _content_type_for(upload, normalized_name)

    if not _is_allowed(content_type, extension):
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type for '{upload.filename}'.",
        )

    data = await _read_file_limited(upload, max_size=settings.upload_max_size_bytes)
    if not data:
        raise HTTPException(status_code=400, detail=f"File '{upload.filename}' is empty.")

    file_id = str(uuid4())
    root = _storage_root()
    stored_name = f"{file_id}{extension}"
    stored_path = root / "files" / stored_name
    stored_path.write_bytes(data)

    preview_text, extraction_note = _extract_text_preview(data, content_type, extension)
    attachment = StoredAttachment(
        id=file_id,
        filename=normalized_name,
        content_type=content_type,
        media_type=_media_type(content_type, extension),
        size_bytes=len(data),
        storage_path=str(stored_path),
        preview_text=preview_text,
        extraction_note=extraction_note,
        created_at=datetime.now(timezone.utc),
    )
    return attachment


def store_generated_artifact(
    *,
    filename: str,
    payload: bytes,
    content_type: str | None = None,
) -> StoredAttachment:
    if not filename:
        raise HTTPException(status_code=400, detail="Generated artifact is missing a filename.")
    if not payload:
        raise HTTPException(status_code=400, detail=f"Generated artifact '{filename}' is empty.")

    normalized_name = _normalize_filename(filename)
    extension = _file_extension(normalized_name)
    resolved_content_type = (
        (content_type or "").strip().lower()
        or (mimetypes.guess_type(normalized_name)[0] or "application/octet-stream").lower()
    )

    file_id = str(uuid4())
    root = _storage_root()
    stored_name = f"{file_id}{extension}"
    stored_path = root / "files" / stored_name
    stored_path.write_bytes(payload)

    preview_text, extraction_note = _extract_text_preview(payload, resolved_content_type, extension)
    return StoredAttachment(
        id=file_id,
        filename=normalized_name,
        content_type=resolved_content_type,
        media_type=_media_type(resolved_content_type, extension),
        size_bytes=len(payload),
        storage_path=str(stored_path),
        preview_text=preview_text,
        extraction_note=extraction_note,
        created_at=datetime.now(timezone.utc),
    )


def load_attachment(file_id: str) -> StoredAttachment | None:
    path = _metadata_path(file_id)
    if not path.exists():
        return None
    payload = json.loads(path.read_text(encoding="utf-8"))
    return StoredAttachment.model_validate(payload)


def from_db_attachment(attachment: Attachment) -> StoredAttachment:
    return StoredAttachment(
        id=str(attachment.id),
        filename=attachment.filename,
        content_type=attachment.content_type,
        media_type=attachment.media_type,
        size_bytes=attachment.size_bytes,
        storage_path=attachment.storage_path,
        preview_text=attachment.preview_text,
        extraction_note=attachment.extraction_note,
        created_at=attachment.created_at,
    )


async def load_attachments_for_ids(
    session: AsyncSession,
    attachment_ids: list[str],
    *,
    allow_json_fallback: bool = True,
) -> list[StoredAttachment]:
    if not attachment_ids:
        return []

    try:
        uuid_ids = [UUID(item) for item in attachment_ids]
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="One or more attachment IDs are invalid.") from exc

    stmt = select(Attachment).where(Attachment.id.in_(uuid_ids))
    rows = (await session.execute(stmt)).scalars().all()
    db_by_id = {str(item.id): from_db_attachment(item) for item in rows}

    loaded: list[StoredAttachment] = []
    missing: list[str] = []
    for file_id in attachment_ids:
        item = db_by_id.get(file_id)
        if item is not None:
            loaded.append(item)
            continue
        if allow_json_fallback:
            fallback = load_attachment(file_id)
            if fallback is not None:
                loaded.append(fallback)
                continue
        missing.append(file_id)

    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Attachment(s) not found: {', '.join(missing)}.",
        )
    return loaded


def _attachment_file_path(attachment: StoredAttachment) -> Path:
    return resolve_storage_path(attachment.storage_path)


def resolve_storage_path(storage_path: str) -> Path:
    path = Path(storage_path)
    if not path.is_absolute():
        project_root = Path(__file__).resolve().parents[3]
        path = project_root / path
    return path


def _encode_attachment_base64(attachment: StoredAttachment) -> str:
    path = _attachment_file_path(attachment)
    if not path.exists():
        raise HTTPException(
            status_code=400,
            detail=f"Attachment '{attachment.id}' file payload was not found on disk.",
        )
    return base64.b64encode(path.read_bytes()).decode("ascii")


def build_attachment_message_parts(attachment_ids: list[str]) -> tuple[str, list[dict[str, Any]]]:
    sections: list[str] = []
    content_blocks: list[dict[str, Any]] = []

    for index, file_id in enumerate(attachment_ids, start=1):
        item = load_attachment(file_id)
        if item is None:
            raise HTTPException(status_code=400, detail=f"Attachment '{file_id}' was not found.")

        if item.media_type == "image":
            # Use the multimodal image block expected by ChatOpenAI / OpenAI-like providers.
            content_blocks.append(
                {
                    "type": "image",
                    "base64": _encode_attachment_base64(item),
                    "mime_type": item.content_type,
                }
            )
            sections.append(
                "\n".join(
                    [
                        f"Attachment {index}",
                        f"- filename: {item.filename}",
                        "- delivery: inline image block",
                    ]
                )
            )
            continue

        if item.media_type == "pdf":
            # OpenAI PDF inputs require explicit filename in the content block.
            content_blocks.append(
                {
                    "type": "file",
                    "base64": _encode_attachment_base64(item),
                    "mime_type": "application/pdf",
                    "filename": item.filename,
                }
            )
            sections.append(
                "\n".join(
                    [
                        f"Attachment {index}",
                        f"- filename: {item.filename}",
                        "- delivery: inline pdf file block",
                    ]
                )
            )
            continue

        lines = [
            f"Attachment {index}",
            f"- filename: {item.filename}",
            f"- media_type: {item.media_type}",
            f"- content_type: {item.content_type}",
            f"- size_bytes: {item.size_bytes}",
        ]
        if item.preview_text:
            lines.append(
                "- extracted_content_preview: This is a preview of the actual content; full content may be larger."
            )
            lines.append(item.preview_text)
        elif item.extraction_note:
            lines.append(f"- extraction_note: {item.extraction_note}")
        else:
            lines.append("- extraction_note: No extracted text.")
        sections.append("\n".join(lines))

    return "\n\n".join(sections), content_blocks


def build_attachment_message_parts_for_items(
    attachments: list[StoredAttachment],
) -> tuple[str, list[dict[str, Any]]]:
    sections: list[str] = []
    content_blocks: list[dict[str, Any]] = []

    for index, item in enumerate(attachments, start=1):
        if item.media_type == "image":
            content_blocks.append(
                {
                    "type": "image",
                    "base64": _encode_attachment_base64(item),
                    "mime_type": item.content_type,
                }
            )
            sections.append(
                "\n".join(
                    [
                        f"Attachment {index}",
                        f"- filename: {item.filename}",
                        "- delivery: inline image block",
                    ]
                )
            )
            continue

        if item.media_type == "pdf":
            content_blocks.append(
                {
                    "type": "file",
                    "base64": _encode_attachment_base64(item),
                    "mime_type": "application/pdf",
                    "filename": item.filename,
                }
            )
            sections.append(
                "\n".join(
                    [
                        f"Attachment {index}",
                        f"- filename: {item.filename}",
                        "- delivery: inline pdf file block",
                    ]
                )
            )
            continue

        lines = [
            f"Attachment {index}",
            f"- filename: {item.filename}",
            f"- media_type: {item.media_type}",
            f"- content_type: {item.content_type}",
            f"- size_bytes: {item.size_bytes}",
        ]
        if item.preview_text:
            lines.append(
                "- (first 5 rows and 20 columns are shown)"
            )
            lines.append(item.preview_text)
        elif item.extraction_note:
            lines.append(f"- extraction_note: {item.extraction_note}")
        else:
            lines.append("- extraction_note: No extracted text.")
        sections.append("\n".join(lines))

    return "\n\n".join(sections), content_blocks


async def build_attachment_message_parts_async(
    session: AsyncSession,
    attachment_ids: list[str],
    *,
    allow_json_fallback: bool = True,
) -> tuple[str, list[dict[str, Any]]]:
    items = await load_attachments_for_ids(
        session,
        attachment_ids,
        allow_json_fallback=allow_json_fallback,
    )
    return build_attachment_message_parts_for_items(items)


def build_attachment_prompt(attachment_ids: list[str]) -> str:
    prompt, _ = build_attachment_message_parts(attachment_ids)
    return prompt
