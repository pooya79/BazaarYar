from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from server.agents.attachments import resolve_storage_path
from server.db.models import Attachment

from .errors import ImportFormatError
from .schema import validate_row_values
from .types import ImportFormat, ReferenceTableColumnInput, TableDataType

_HEADER_TOKEN_RE = re.compile(r"[^a-zA-Z0-9_]+")


@dataclass
class ParsedImportData:
    source_format: ImportFormat
    rows: list[dict[str, Any]]


def _normalize_header(value: str, *, fallback_index: int) -> str:
    candidate = _HEADER_TOKEN_RE.sub("_", value.strip()).strip("_")
    if not candidate:
        candidate = f"column_{fallback_index}"
    if not re.match(r"^[a-zA-Z_]", candidate):
        candidate = f"column_{candidate}"
    return candidate[:63]


def _dedupe_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    output: list[str] = []
    for header in headers:
        count = seen.get(header, 0)
        if count:
            deduped = f"{header}_{count + 1}"
        else:
            deduped = header
        seen[header] = count + 1
        output.append(deduped[:63])
    return output


def _parse_csv(
    payload: bytes,
    *,
    has_header: bool,
    delimiter: str | None,
) -> list[dict[str, Any]]:
    decoded = payload.decode("utf-8", errors="replace")
    if delimiter is None:
        try:
            dialect = csv.Sniffer().sniff(decoded[:2048])
            resolved_delimiter = dialect.delimiter
        except Exception:
            resolved_delimiter = ","
    else:
        resolved_delimiter = delimiter

    reader = csv.reader(io.StringIO(decoded), delimiter=resolved_delimiter)
    rows = list(reader)
    if not rows:
        return []

    if has_header:
        raw_headers = rows[0]
        headers = _dedupe_headers(
            [
                _normalize_header(str(value), fallback_index=index + 1)
                for index, value in enumerate(raw_headers)
            ]
        )
        data_rows = rows[1:]
    else:
        width = max((len(row) for row in rows), default=0)
        headers = [f"column_{index + 1}" for index in range(width)]
        data_rows = rows

    output: list[dict[str, Any]] = []
    for row in data_rows:
        item: dict[str, Any] = {}
        for index, header in enumerate(headers):
            item[header] = row[index] if index < len(row) else None
        output.append(item)
    return output


def _parse_json(payload: bytes) -> list[dict[str, Any]]:
    parsed = json.loads(payload.decode("utf-8", errors="replace"))
    if isinstance(parsed, dict) and isinstance(parsed.get("rows"), list):
        parsed = parsed["rows"]

    if not isinstance(parsed, list):
        raise ImportFormatError("JSON import expects a list of objects or {\"rows\": [...]}.")

    output: list[dict[str, Any]] = []
    for index, item in enumerate(parsed, start=1):
        if not isinstance(item, dict):
            raise ImportFormatError(f"JSON row {index} is not an object.")
        normalized: dict[str, Any] = {}
        for key, value in item.items():
            normalized[_normalize_header(str(key), fallback_index=1)] = value
        output.append(normalized)
    return output


def _parse_xlsx(payload: bytes, *, has_header: bool) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return _parse_xlsx_stdlib(payload, has_header=has_header)

    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook.worksheets[0] if workbook.worksheets else None
    if sheet is None:
        return []

    rows: list[list[Any]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([cell for cell in row])
    return _rows_to_objects(rows, has_header=has_header)


def _parse_xlsx_stdlib(payload: bytes, *, has_header: bool) -> list[dict[str, Any]]:
    ns = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with zipfile.ZipFile(io.BytesIO(payload)) as zf:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            shared_root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for item in shared_root.findall(".//x:si", ns):
                parts = [node.text or "" for node in item.findall(".//x:t", ns)]
                shared_strings.append("".join(parts))

        sheet_names = sorted(
            name
            for name in zf.namelist()
            if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        )
        if not sheet_names:
            return []

        sheet_root = ET.fromstring(zf.read(sheet_names[0]))
        rows: list[list[str | None]] = []
        for row in sheet_root.findall(".//x:sheetData/x:row", ns):
            values: list[str | None] = []
            for cell in row.findall("x:c", ns):
                cell_type = cell.attrib.get("t")
                inline_node = cell.find("x:is/x:t", ns)
                value_node = cell.find("x:v", ns)
                if inline_node is not None and inline_node.text is not None:
                    values.append(inline_node.text)
                    continue
                if value_node is None or value_node.text is None:
                    values.append(None)
                    continue
                raw = value_node.text
                if cell_type == "s":
                    try:
                        values.append(shared_strings[int(raw)])
                    except Exception:
                        values.append(raw)
                else:
                    values.append(raw)
            rows.append(values)

    return _rows_to_objects(rows, has_header=has_header)


def _rows_to_objects(rows: list[list[Any]], *, has_header: bool) -> list[dict[str, Any]]:
    if not rows:
        return []

    if has_header:
        headers = _dedupe_headers(
            [
                _normalize_header(str(cell or ""), fallback_index=index + 1)
                for index, cell in enumerate(rows[0])
            ]
        )
        data_rows = rows[1:]
    else:
        width = max((len(row) for row in rows), default=0)
        headers = [f"column_{index + 1}" for index in range(width)]
        data_rows = rows

    output: list[dict[str, Any]] = []
    for row in data_rows:
        item: dict[str, Any] = {}
        for index, header in enumerate(headers):
            item[header] = row[index] if index < len(row) else None
        output.append(item)
    return output


def detect_format(
    *,
    source_format: ImportFormat | None,
    filename: str | None,
) -> ImportFormat:
    if source_format is not None:
        return source_format

    suffix = Path(filename or "").suffix.lower()
    if suffix == ".csv":
        return ImportFormat.CSV
    if suffix == ".json":
        return ImportFormat.JSON
    if suffix in {".xlsx", ".xlsm"}:
        return ImportFormat.XLSX

    raise ImportFormatError("Could not infer import format. Provide source_format explicitly.")


def parse_attachment(
    attachment: Attachment,
    *,
    source_format: ImportFormat | None,
    has_header: bool,
    delimiter: str | None,
) -> ParsedImportData:
    path = resolve_storage_path(attachment.storage_path)
    if not path.exists():
        raise ImportFormatError(f"Attachment file '{attachment.id}' is missing on disk.")

    resolved_format = detect_format(source_format=source_format, filename=attachment.filename)
    payload = path.read_bytes()

    if resolved_format == ImportFormat.CSV:
        rows = _parse_csv(payload, has_header=has_header, delimiter=delimiter)
    elif resolved_format == ImportFormat.JSON:
        rows = _parse_json(payload)
    elif resolved_format == ImportFormat.XLSX:
        rows = _parse_xlsx(payload, has_header=has_header)
    else:  # pragma: no cover - enum guard
        raise ImportFormatError(f"Unsupported format '{resolved_format}'.")

    return ParsedImportData(source_format=resolved_format, rows=rows)


def _infer_type(values: list[Any]) -> tuple[TableDataType, float]:
    non_null = [item for item in values if item not in (None, "")]
    if not non_null:
        return TableDataType.TEXT, 0.4

    def _all(pred) -> bool:
        for value in non_null:
            if not pred(value):
                return False
        return True

    if _all(lambda item: isinstance(item, bool) or str(item).strip().lower() in {"true", "false", "0", "1"}):
        return TableDataType.BOOLEAN, 0.9

    if _all(lambda item: isinstance(item, int) or (isinstance(item, str) and item.strip().isdigit())):
        return TableDataType.INTEGER, 0.9

    def _is_float_like(item: Any) -> bool:
        if isinstance(item, (int, float)):
            return True
        if not isinstance(item, str):
            return False
        try:
            float(item.strip())
            return True
        except Exception:
            return False

    if _all(_is_float_like):
        return TableDataType.FLOAT, 0.85

    def _is_date_like(item: Any) -> bool:
        if isinstance(item, date) and not isinstance(item, datetime):
            return True
        if not isinstance(item, str):
            return False
        try:
            date.fromisoformat(item.strip())
            return True
        except Exception:
            return False

    if _all(_is_date_like):
        return TableDataType.DATE, 0.8

    def _is_timestamp_like(item: Any) -> bool:
        if isinstance(item, datetime):
            return True
        if not isinstance(item, str):
            return False
        try:
            datetime.fromisoformat(item.strip().replace("Z", "+00:00"))
            return True
        except Exception:
            return False

    if _all(_is_timestamp_like):
        return TableDataType.TIMESTAMP, 0.75

    if _all(lambda item: isinstance(item, (dict, list))):
        return TableDataType.JSON, 0.7

    return TableDataType.TEXT, 0.6


def infer_columns(
    rows: list[dict[str, Any]],
    *,
    max_columns: int,
) -> list[dict[str, Any]]:
    if not rows:
        return []

    all_keys: list[str] = []
    for row in rows:
        for key in row.keys():
            if key not in all_keys:
                all_keys.append(key)

    if len(all_keys) > max_columns:
        raise ImportFormatError(f"Too many columns in source file. Maximum is {max_columns}.")

    inferred: list[dict[str, Any]] = []
    for key in all_keys:
        sample = [row.get(key) for row in rows[:250]]
        inferred_type, confidence = _infer_type(sample)
        inferred.append(
            {
                "name": key,
                "data_type": inferred_type.value,
                "confidence": confidence,
            }
        )
    return inferred


def validate_import_rows(
    rows: list[dict[str, Any]],
    *,
    columns: list[ReferenceTableColumnInput],
    max_rows: int,
    max_cell_length: int,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    if len(rows) > max_rows:
        raise ImportFormatError(f"Import row count exceeds maximum of {max_rows}.")

    normalized_rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []

    for index, row in enumerate(rows, start=1):
        try:
            normalized = validate_row_values(
                row,
                columns=columns,
                max_cell_length=max_cell_length,
            )
            normalized_rows.append(normalized)
        except Exception as exc:
            if hasattr(exc, "errors"):
                errors.append(
                    {
                        "row": index,
                        "error": str(exc),
                        "details": getattr(exc, "errors", []),
                    }
                )
            else:
                errors.append({"row": index, "error": str(exc)})

    return normalized_rows, errors
